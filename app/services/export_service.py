# -*- coding: utf-8 -*-
"""تصدير التقارير المالية: Excel (openpyxl) و PDF (reportlab) بدعم كامل للغة العربية"""
import io

import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_FONT_PATH = "app/static/fonts/Cairo-Regular.ttf"

pdfmetrics.registerFont(TTFont("Cairo", _FONT_PATH))

_HDR_BG = "1E3A5F"
_SUB_BG = "E8EEF4"
_ACC_BG = "DCE7F1"


def ar(text) -> str:
    """إعادة تشكيل النص العربي وعكس اتجاهه لعرضه بشكل صحيح داخل PDF"""
    return get_display(arabic_reshaper.reshape(str(text)))


def money(value) -> str:
    try:
        return f"{float(value or 0):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


# --------------------------------------------------------------------------
# Excel (openpyxl)
# --------------------------------------------------------------------------
def _xlsx_file(sheet_title: str, headers: list[str], widths: list[float],
               body: list[list]) -> bytes:
    """ينشئ ملف Excel: صف عناوين منمق + صفوف بيانات"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    header_fill = PatternFill("solid", fgColor=_HDR_BG)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + col)].width = w
    for r in body:
        ws.append(r)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def chart_excel(accounts: list[dict]) -> bytes:
    headers = ["الكود", "اسم الحساب", "النوع", "المستوى", "الرصيد (ج.م)"]
    widths = [14, 55, 16, 10, 16]
    body = [[a["code"], a["name"], a["type"], a["level"], money(a["balance"])]
            for a in accounts]
    return _xlsx_file("شجرة الحسابات", headers, widths, body)


def journal_excel(rows: list[dict]) -> bytes:
    headers = ["رقم المرجع", "التاريخ", "البيان", "الحساب", "مدين (ج.م)", "دائن (ج.م)"]
    widths = [12, 12, 45, 30, 14, 14]
    body = []
    for r in rows:
        body.append([
            f"#{r['entry_id']}", str(r["date"]), r["description"],
            f"{r['account_code']} - {r['account_name']}",
            money(r["debit"]) if r["debit"] else "—",
            money(r["credit"]) if r["credit"] else "—",
        ])
    return _xlsx_file("اليومية الأمريكية", headers, widths, body)


def ledger_excel(blocks: list[dict]) -> bytes:
    headers = ["كود الحساب", "الحساب", "التاريخ", "رقم المرجع", "البيان",
               "مدين (ج.م)", "دائن (ج.م)", "الرصيد الجاري"]
    widths = [12, 30, 12, 12, 40, 14, 14, 14]
    body = []
    for b in blocks:
        body.append([b["code"], b["name"] + f" (فتح: {money(b['opening'])})", "", "", "", "", "", ""])
        for ln in b["lines"]:
            body.append([
                "", "", str(ln["date"]), f"#{ln['entry_id']}", ln["description"],
                money(ln["debit"]) if ln["debit"] else "—",
                money(ln["credit"]) if ln["credit"] else "—",
                money(ln["running"]),
            ])
        body.append(["", f"إجمالي {b['name']} (رصيد ختامي)", "", "", "", "", "",
                     money(b["closing"])])
    return _xlsx_file("الأستاذ العام", headers, widths, body)


# --------------------------------------------------------------------------
# PDF (reportlab) — RTL support
# --------------------------------------------------------------------------
def _rtl_para(text, style):
    """إنشاء Paragraph عربي مع معالجة الاتجاه الصحيح"""
    return Paragraph(ar(text), style)


def _pdf_file(title: str, headers: list[str], col_widths_mm: list[float],
              body: list[tuple[list, str]], landscape_page: bool = True) -> bytes:
    """body: قائمة (صف، kind) حيث kind ∈ {row, sub, acc} للتمييز بالألوان"""
    page = landscape(A4) if landscape_page else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=page,
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title=ar(title))

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleAr", parent=styles["Title"],
                                 fontName="Cairo", fontSize=16, leading=22,
                                 alignment=TA_CENTER,
                                 textColor=colors.HexColor("#" + _HDR_BG))
    head_style = ParagraphStyle("HeadAr", parent=styles["Normal"],
                                fontName="Cairo", fontSize=9, leading=12,
                                textColor=colors.white, alignment=TA_CENTER)
    cell_style = ParagraphStyle("CellAr", parent=styles["Normal"],
                                fontName="Cairo", fontSize=8, leading=11,
                                alignment=TA_RIGHT)
    cell_center = ParagraphStyle("CellC", parent=cell_style, alignment=TA_CENTER)
    cell_left = ParagraphStyle("CellL", parent=cell_style, alignment=TA_LEFT)
    sub_style = ParagraphStyle("SubAr", parent=cell_style, fontName="Cairo",
                               fontSize=8, leading=11, alignment=TA_CENTER,
                               textColor=colors.black)

    story = [_rtl_para(title, title_style), Spacer(1, 4 * mm)]

    # عكس ترتيب الأعمدة RTL → الأعمدة تظهر من اليمين لليسار
    rtl_headers = list(reversed(headers))

    data = [[_rtl_para(h, head_style) for h in rtl_headers]]
    styled_rows = []
    for row, kind in body:
        rtl_row = list(reversed(row))
        if kind == "sub":
            data.append([_rtl_para(v, sub_style) for v in rtl_row])
        else:
            data.append([_rtl_para(v, cell_style) for v in rtl_row])
        styled_rows.append((len(data) - 1, kind))

    table = Table(data, colWidths=[w * mm for w in reversed(col_widths_mm)],
                  repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + _HDR_BG)),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B9C4CF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for row_idx, kind in styled_rows:
        if kind == "acc":
            style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx),
                               colors.HexColor("#" + _ACC_BG)))
        elif kind == "sub":
            style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx),
                               colors.HexColor("#" + _SUB_BG)))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def chart_pdf(accounts: list[dict]) -> bytes:
    headers = ["الكود", "اسم الحساب", "النوع", "المستوى", "الرصيد (ج.م)"]
    body = [([a["code"], a["name"], a["type"], str(a["level"]), money(a["balance"])], "row")
            for a in accounts]
    return _pdf_file("شجرة دليل الحسابات", headers, [20, 70, 30, 20, 30], body,
                     landscape_page=False)


def journal_pdf(rows: list[dict]) -> bytes:
    headers = ["رقم المرجع", "التاريخ", "البيان", "الحساب", "مدين (ج.م)", "دائن (ج.م)"]
    body = []
    for r in rows:
        body.append(([
            f"#{r['entry_id']}", str(r["date"]), r["description"],
            f"{r['account_code']} - {r['account_name']}",
            money(r["debit"]) if r["debit"] else "—",
            money(r["credit"]) if r["credit"] else "—",
        ], "row"))
    return _pdf_file("اليومية الأمريكية", headers, [20, 22, 55, 55, 22, 22], body)


def ledger_pdf(blocks: list[dict]) -> bytes:
    headers = ["الحساب", "التاريخ", "رقم المرجع", "البيان", "مدين (ج.م)", "دائن (ج.م)", "الرصيد"]
    body = []
    for b in blocks:
        body.append(([f"{b['code']} - {b['name']}   (رصيد افتتاحي: {money(b['opening'])})",
                      "", "", "", "", "", ""], "acc"))
        for ln in b["lines"]:
            body.append(([
                "", str(ln["date"]), f"#{ln['entry_id']}", ln["description"],
                money(ln["debit"]) if ln["debit"] else "—",
                money(ln["credit"]) if ln["credit"] else "—",
                money(ln["running"]),
            ], "row"))
        body.append((["", "", "", f"إجمالي الحساب - الرصيد الختامي", "", "",
                      money(b["closing"])], "sub"))
    return _pdf_file("الأستاذ العام", headers, [55, 20, 18, 50, 20, 20, 20], body)
